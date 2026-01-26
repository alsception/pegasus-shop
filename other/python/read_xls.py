from openpyxl import load_workbook

file_path = "/home/nix/Downloads/Rezervacije Barbacoa 2025-12.xlsx"

workbook = load_workbook(filename=file_path, data_only=True)

for sheet in workbook.worksheets:
    print("-------------------------------------------------")
    print(f"Sheet: {sheet.title}")
    print("-------------------------------------------------")

    for row in sheet.iter_rows(values_only=True):
        # row[1] = druga kolona (index 1)
        second_col = row[1] if row and len(row) > 1 else None

        # preskoci ako je prazno
        if second_col is None or str(second_col).strip() == "":
            continue

        # preskoci ako sadrzi 'Ime i Prezime'
        if "Ime i Prezime" in str(second_col):
            continue

        print(row)
