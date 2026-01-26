from openpyxl import load_workbook
from datetime import time, datetime

file_path = "/home/nix/Downloads/Rezervacije Barbacoa 2025-12.xlsx"

workbook = load_workbook(filename=file_path, data_only=True)

def convert_value(v):
    """Convert datetime/time to string"""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    return v  # leave everything else as is

for sheet in workbook.worksheets:
    print("-------------------------------------------------")
    print(f"Sheet: {sheet.title}")
    print("-------------------------------------------------")

    for row in sheet.iter_rows(values_only=True):
        # row[1] = second column
        second_col = row[1] if row and len(row) > 1 else None

        # skip empty
        if second_col is None or str(second_col).strip() == "":
            continue

        # skip header
        if "Ime i Prezime" in str(second_col):
            continue

        # convert all values in the row
        converted_row = tuple(convert_value(v) for v in row)

        print(converted_row)

